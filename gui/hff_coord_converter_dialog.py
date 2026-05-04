# -*- coding: utf-8 -*-
"""HFF Coordinate Converter dialog.

Convert coordinates between DDM (Garmin), DMS, decimal degrees and UTM
using the shared pure module ``modules.utility.coord_converter``.

Sections:
  1. Single conversion (lat/lon → all formats).
  2. Batch table — collect multiple points across conversions, import
     from CSV/Excel, export to CSV/Excel.
  3. GIS export — write the batch (or current single point) to a
     GeoPackage at a chosen EPSG and load it into the QGIS TOC.

Built with ``qgis.PyQt`` so it works on both Qt5 (current QGIS LTR)
and Qt6 (QGIS 3.99+/QGIS 4) without source changes.
"""
from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path

from qgis.PyQt.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from modules.utility.coord_converter import (
    HAS_PYPROJ,
    convert_all,
    parse_coordinate,
    utm_zone_from_lon,
)

try:
    import openpyxl  # type: ignore
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


_BATCH_COLS = ["Name", "Lat (DD)", "Lon (DD)", "UTM zone", "X", "Y"]


def _hff_home() -> Path:
    return Path(os.environ.get("HFF_HOME", str(Path.home() / "HFF")))


class CoordConverterDialog(QDialog):
    """DDM / DMS / DD / UTM converter + batch + GeoPackage export."""

    def __init__(self, parent=None, iface=None):
        super().__init__(parent)
        self.setWindowTitle("HFF — Coordinate Converter")
        self.resize(820, 720)
        self._iface = iface
        self._outputs: dict[str, QLineEdit] = {}
        self._last_result: dict | None = None
        self._build_ui()
        self.lat_input.setText("N 34 01.825")
        self.lon_input.setText("E 035 37.349")
        self._convert()

    # ------------------------------------------------------------------ UI

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)

        tabs = QTabWidget()
        root.addWidget(tabs)

        tabs.addTab(self._build_convert_tab(), "Convert")
        tabs.addTab(self._build_batch_tab(), "Batch")
        tabs.addTab(self._build_gis_tab(), "GIS export")

    def _build_convert_tab(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page)

        # ---- Input ----
        in_box = QGroupBox("Input")
        form = QFormLayout()
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Optional point name (used in Batch + GeoPackage)")
        self.lat_input = QLineEdit()
        self.lon_input = QLineEdit()
        form.addRow("Name:", self.name_input)
        form.addRow("Latitude:", self.lat_input)
        form.addRow("Longitude:", self.lon_input)
        info = QLabel(
            "Accepted formats: <code>N 34 01.825</code> (DDM Garmin) · "
            "<code>34°1'49.5\"N</code> (DMS) · <code>34.030417</code> (DD)"
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: gray;")
        form.addRow(info)
        in_box.setLayout(form)
        root.addWidget(in_box)

        # ---- UTM options ----
        utm_box = QGroupBox("UTM options")
        utm_h = QHBoxLayout()
        utm_h.addWidget(QLabel("Zone:"))
        self.zone_combo = QComboBox()
        self.zone_combo.addItem("Auto")
        for z in range(1, 61):
            self.zone_combo.addItem(str(z))
        utm_h.addWidget(self.zone_combo)
        engine_label = QLabel(
            f"  Engine: {'pyproj' if HAS_PYPROJ else 'pure-python (no pyproj)'}"
        )
        engine_label.setStyleSheet("color: gray;")
        utm_h.addWidget(engine_label)
        utm_h.addStretch(1)
        utm_box.setLayout(utm_h)
        root.addWidget(utm_box)

        # ---- Buttons ----
        btn_row = QHBoxLayout()
        convert_btn = QPushButton("Convert")
        clear_btn = QPushButton("Clear")
        copy_dd_btn = QPushButton("Copy DD")
        copy_utm_btn = QPushButton("Copy UTM (X, Y)")
        add_batch_btn = QPushButton("➕ Add to batch")
        btn_row.addWidget(convert_btn)
        btn_row.addWidget(clear_btn)
        btn_row.addWidget(copy_dd_btn)
        btn_row.addWidget(copy_utm_btn)
        btn_row.addWidget(add_batch_btn)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

        # ---- Output ----
        out_box = QGroupBox("Results")
        out_form = QFormLayout()
        for label, key in [
            ("DD (lat, lon):", "dd"),
            ("DDM (Garmin):", "ddm"),
            ("DMS:", "dms"),
            ("GIS (lon, lat):", "gis"),
            ("UTM:", "utm_text"),
            ("EPSG:", "epsg_text"),
            ("Google Maps:", "gmaps_url"),
        ]:
            entry = QLineEdit()
            entry.setReadOnly(True)
            self._outputs[key] = entry
            out_form.addRow(label, entry)
        out_box.setLayout(out_form)
        root.addWidget(out_box)
        root.addStretch(1)

        # ---- Wiring ----
        convert_btn.clicked.connect(self._convert)
        clear_btn.clicked.connect(self._clear_single)
        copy_dd_btn.clicked.connect(self._copy_dd)
        copy_utm_btn.clicked.connect(self._copy_utm)
        add_batch_btn.clicked.connect(self._add_to_batch)
        self.lat_input.returnPressed.connect(self._convert)
        self.lon_input.returnPressed.connect(self._convert)
        self.zone_combo.currentIndexChanged.connect(self._convert)

        return page

    def _build_batch_tab(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page)

        info = QLabel(
            "Collect multiple points across conversions, import from "
            "CSV / Excel, export to CSV / Excel. Coordinates in any "
            "input format are auto-converted to DD."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: gray;")
        root.addWidget(info)

        self.batch_table = QTableWidget(0, len(_BATCH_COLS))
        self.batch_table.setHorizontalHeaderLabels(_BATCH_COLS)
        # Use fully-qualified enum form for Qt5/Qt6 compatibility.
        try:
            sel_rows = QAbstractItemView.SelectionBehavior.SelectRows
        except AttributeError:
            sel_rows = QAbstractItemView.SelectRows
        self.batch_table.setSelectionBehavior(sel_rows)
        try:
            stretch = QHeaderView.ResizeMode.Stretch
        except AttributeError:
            stretch = QHeaderView.Stretch
        self.batch_table.horizontalHeader().setSectionResizeMode(stretch)
        root.addWidget(self.batch_table)

        # Buttons
        btn_row = QHBoxLayout()
        remove_btn = QPushButton("Remove selected")
        clear_btn = QPushButton("Clear all")
        import_btn = QPushButton("Import CSV / Excel…")
        export_csv_btn = QPushButton("Export CSV…")
        export_xlsx_btn = QPushButton("Export Excel…")
        if not HAS_OPENPYXL:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip(
                "Install openpyxl to enable Excel export"
            )
        btn_row.addWidget(remove_btn)
        btn_row.addWidget(clear_btn)
        btn_row.addWidget(import_btn)
        btn_row.addWidget(export_csv_btn)
        btn_row.addWidget(export_xlsx_btn)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

        remove_btn.clicked.connect(self._remove_selected_batch)
        clear_btn.clicked.connect(self._clear_batch)
        import_btn.clicked.connect(self._import_batch)
        export_csv_btn.clicked.connect(self._export_csv)
        export_xlsx_btn.clicked.connect(self._export_xlsx)

        return page

    def _build_gis_tab(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page)

        info = QLabel(
            "Write the batch table (or the current single point if "
            "batch is empty) to a GeoPackage at the chosen CRS and "
            "load the resulting layer into the QGIS Layers panel."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: gray;")
        root.addWidget(info)

        # CRS choice
        crs_box = QGroupBox("Target CRS")
        crs_form = QFormLayout()
        self.crs_combo = QComboBox()
        self.crs_combo.addItem("WGS84 (EPSG:4326)", "epsg:4326")
        self.crs_combo.addItem("UTM auto-zone (WGS84)", "utm:auto")
        self.crs_combo.addItem("Use current QGIS project CRS", "project")
        self.crs_combo.addItem("Custom EPSG…", "custom")
        crs_form.addRow("CRS:", self.crs_combo)
        self.custom_epsg_input = QLineEdit()
        self.custom_epsg_input.setPlaceholderText("e.g. 32636")
        self.custom_epsg_input.setEnabled(False)
        crs_form.addRow("Custom EPSG number:", self.custom_epsg_input)
        crs_box.setLayout(crs_form)
        root.addWidget(crs_box)

        self.crs_combo.currentIndexChanged.connect(
            lambda _: self.custom_epsg_input.setEnabled(
                self.crs_combo.currentData() == "custom"
            )
        )

        # Output path
        out_box = QGroupBox("Output GeoPackage")
        out_form = QFormLayout()
        out_h = QHBoxLayout()
        self.gpkg_path_input = QLineEdit()
        self.gpkg_path_input.setText(self._default_gpkg_path())
        browse_btn = QPushButton("Browse…")
        browse_btn.clicked.connect(self._browse_gpkg)
        out_h.addWidget(self.gpkg_path_input)
        out_h.addWidget(browse_btn)
        out_form.addRow("Path:", out_h)
        self.layer_name_input = QLineEdit()
        self.layer_name_input.setText(self._default_layer_name())
        out_form.addRow("Layer name:", self.layer_name_input)
        out_box.setLayout(out_form)
        root.addWidget(out_box)

        # Action button
        export_btn = QPushButton("Export to GeoPackage and load layer")
        export_btn.clicked.connect(self._export_gpkg)
        root.addWidget(export_btn)

        root.addStretch(1)
        return page

    # ------------------------------------------------------------------ Convert tab actions

    def _convert(self) -> None:
        lat_text = self.lat_input.text()
        lon_text = self.lon_input.text()
        if not lat_text.strip() or not lon_text.strip():
            return
        zone_sel = self.zone_combo.currentText()
        forced_zone = None if zone_sel == "Auto" else int(zone_sel)
        try:
            r = convert_all(lat_text, lon_text, zone=forced_zone)
        except ValueError as e:
            QMessageBox.warning(self, "Invalid coordinate", str(e))
            return
        self._last_result = r

        utm_extra = ""
        if forced_zone is not None and forced_zone != r["auto_zone"]:
            utm_extra = f"  (auto zone: {r['auto_zone']}{r['utm_hemi']})"
        for key in ("dd", "ddm", "dms", "gis", "epsg_text", "gmaps_url"):
            self._outputs[key].setText(r[key])
        self._outputs["utm_text"].setText(r["utm_text"] + utm_extra)

    def _clear_single(self) -> None:
        self.name_input.clear()
        self.lat_input.clear()
        self.lon_input.clear()
        for entry in self._outputs.values():
            entry.clear()
        self._last_result = None
        self.lat_input.setFocus()

    def _copy_dd(self) -> None:
        if self._last_result is None:
            return
        QApplication.clipboard().setText(self._last_result["dd"])

    def _copy_utm(self) -> None:
        if self._last_result is None:
            return
        text = (
            f"{self._last_result['utm_easting']:.0f}, "
            f"{self._last_result['utm_northing']:.0f}"
        )
        QApplication.clipboard().setText(text)

    def _add_to_batch(self) -> None:
        if self._last_result is None:
            QMessageBox.information(
                self, "Nothing to add", "Convert a point first."
            )
            return
        name = self.name_input.text().strip()
        if not name:
            name = f"point_{self.batch_table.rowCount() + 1}"
        self._append_batch_row(
            name=name,
            lat_dd=self._last_result["lat_dd"],
            lon_dd=self._last_result["lon_dd"],
            utm_zone=f"{self._last_result['utm_zone']}{self._last_result['utm_hemi']}",
            x=self._last_result["utm_easting"],
            y=self._last_result["utm_northing"],
        )

    # ------------------------------------------------------------------ Batch tab actions

    def _append_batch_row(self, *, name: str, lat_dd: float, lon_dd: float,
                          utm_zone: str, x: float, y: float) -> None:
        row = self.batch_table.rowCount()
        self.batch_table.insertRow(row)
        values = [
            name,
            f"{lat_dd:.6f}",
            f"{lon_dd:.6f}",
            utm_zone,
            f"{x:.2f}",
            f"{y:.2f}",
        ]
        for col, val in enumerate(values):
            self.batch_table.setItem(row, col, QTableWidgetItem(val))

    def _remove_selected_batch(self) -> None:
        rows = sorted(
            {idx.row() for idx in self.batch_table.selectedIndexes()},
            reverse=True,
        )
        for r in rows:
            self.batch_table.removeRow(r)

    def _clear_batch(self) -> None:
        self.batch_table.setRowCount(0)

    def _read_batch_dd_rows(self) -> list[dict]:
        """Snapshot the table as list of {name, lat, lon}."""
        rows: list[dict] = []
        for r in range(self.batch_table.rowCount()):
            try:
                rows.append({
                    "name": self.batch_table.item(r, 0).text(),
                    "lat": float(self.batch_table.item(r, 1).text()),
                    "lon": float(self.batch_table.item(r, 2).text()),
                })
            except (AttributeError, ValueError):
                continue
        return rows

    # -- Import --------------------------------------------------------

    def _import_batch(self) -> None:
        filt = "Tabular files (*.csv *.xlsx)"
        if not HAS_OPENPYXL:
            filt = "CSV files (*.csv)"
        path, _ = QFileDialog.getOpenFileName(
            self, "Import points", str(_hff_home()), filt
        )
        if not path:
            return
        try:
            if path.lower().endswith((".xlsx", ".xls")):
                rows = self._read_xlsx(path)
            else:
                rows = self._read_csv(path)
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))
            return
        added, skipped = self._add_imported_rows(rows)
        QMessageBox.information(
            self, "Import complete",
            f"Added {added} point(s).\nSkipped {skipped} unparseable row(s)."
        )

    @staticmethod
    def _read_csv(path: str) -> list[dict]:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            return [dict(r) for r in reader]

    @staticmethod
    def _read_xlsx(path: str) -> list[dict]:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out: list[dict] = []
        for raw in rows[1:]:
            d = {}
            for h, v in zip(headers, raw):
                d[h] = "" if v is None else str(v)
            out.append(d)
        return out

    def _add_imported_rows(self, rows: list[dict]) -> tuple[int, int]:
        """Auto-detect lat/lon columns; convert to DD and append.
        Returns (added, skipped)."""
        if not rows:
            return 0, 0
        headers = [k.strip() for k in rows[0].keys()]
        lower = [h.lower() for h in headers]

        def _find(*candidates: str) -> str | None:
            for c in candidates:
                for orig, low in zip(headers, lower):
                    if low == c:
                        return orig
            return None

        lat_col = _find("lat", "latitude", "y", "lat (dd)", "lat_dd")
        lon_col = _find("lon", "lng", "long", "longitude", "x", "lon (dd)", "lon_dd")
        name_col = _find("name", "label", "id", "point", "site")

        if not lat_col or not lon_col:
            raise ValueError(
                "Could not find latitude / longitude columns. "
                "Expected headers like 'lat'/'lon' or 'latitude'/'longitude'."
            )

        added = skipped = 0
        for i, row in enumerate(rows, start=1):
            try:
                lat_dd = parse_coordinate(str(row[lat_col]))
                lon_dd = parse_coordinate(str(row[lon_col]))
                if not -90 <= lat_dd <= 90 or not -180 <= lon_dd <= 180:
                    raise ValueError("out of range")
            except Exception:
                skipped += 1
                continue
            zone = utm_zone_from_lon(lon_dd)
            from modules.utility.coord_converter import latlon_to_utm
            x, y, used_zone, hemi = latlon_to_utm(lat_dd, lon_dd, zone)
            name = str(row.get(name_col) or "").strip() if name_col else ""
            if not name:
                name = f"point_{self.batch_table.rowCount() + 1}"
            self._append_batch_row(
                name=name, lat_dd=lat_dd, lon_dd=lon_dd,
                utm_zone=f"{used_zone}{hemi}", x=x, y=y,
            )
            added += 1
        return added, skipped

    # -- Export CSV/Excel ---------------------------------------------

    def _export_csv(self) -> None:
        if self.batch_table.rowCount() == 0:
            QMessageBox.information(self, "Empty batch", "No rows to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export batch as CSV",
            str(_hff_home() / "coord_batch.csv"),
            "CSV files (*.csv)",
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(_BATCH_COLS)
                for r in range(self.batch_table.rowCount()):
                    w.writerow([
                        self.batch_table.item(r, c).text()
                        for c in range(len(_BATCH_COLS))
                    ])
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        QMessageBox.information(self, "CSV exported", path)

    def _export_xlsx(self) -> None:
        if not HAS_OPENPYXL:
            return
        if self.batch_table.rowCount() == 0:
            QMessageBox.information(self, "Empty batch", "No rows to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export batch as Excel",
            str(_hff_home() / "coord_batch.xlsx"),
            "Excel files (*.xlsx)",
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "coords"
            ws.append(_BATCH_COLS)
            for r in range(self.batch_table.rowCount()):
                ws.append([
                    self.batch_table.item(r, c).text()
                    for c in range(len(_BATCH_COLS))
                ])
            wb.save(path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        QMessageBox.information(self, "Excel exported", path)

    # ------------------------------------------------------------------ GIS export

    def _default_gpkg_path(self) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return str(_hff_home() / f"coord_export_{ts}.gpkg")

    def _default_layer_name(self) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"coord_export_{ts}"

    def _browse_gpkg(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Output GeoPackage",
            self.gpkg_path_input.text() or self._default_gpkg_path(),
            "GeoPackage (*.gpkg)",
        )
        if path:
            self.gpkg_path_input.setText(path)

    def _resolve_target_epsg(self, points: list[dict]) -> int | None:
        choice = self.crs_combo.currentData()
        if choice == "epsg:4326":
            return 4326
        if choice == "utm:auto":
            if not points:
                return 4326
            # Use first point's lon to pick a zone
            lon = points[0]["lon"]
            lat = points[0]["lat"]
            zone = utm_zone_from_lon(lon)
            return (32600 + zone) if lat >= 0 else (32700 + zone)
        if choice == "project":
            try:
                from qgis.core import QgsProject
                code = QgsProject.instance().crs().authid()
                # authid like 'EPSG:32636'
                if code and code.upper().startswith("EPSG:"):
                    return int(code.split(":", 1)[1])
            except Exception:
                pass
            return 4326
        if choice == "custom":
            t = self.custom_epsg_input.text().strip()
            if not t.isdigit():
                QMessageBox.warning(
                    self, "Invalid EPSG",
                    "Type a numeric EPSG code (e.g. 32636).",
                )
                return None
            return int(t)
        return 4326

    def _export_gpkg(self) -> None:
        # Collect points: batch first, otherwise the single result.
        points = self._read_batch_dd_rows()
        if not points and self._last_result is not None:
            name = self.name_input.text().strip() or "point_1"
            points = [{
                "name": name,
                "lat": self._last_result["lat_dd"],
                "lon": self._last_result["lon_dd"],
            }]
        if not points:
            QMessageBox.information(
                self, "No points",
                "Add points to the batch (or convert a single point) first.",
            )
            return

        epsg = self._resolve_target_epsg(points)
        if epsg is None:
            return

        path = self.gpkg_path_input.text().strip()
        if not path:
            path = self._default_gpkg_path()
            self.gpkg_path_input.setText(path)
        if not path.lower().endswith(".gpkg"):
            path += ".gpkg"

        layer_name = self.layer_name_input.text().strip() or self._default_layer_name()

        try:
            self._write_gpkg_and_load(points, path, layer_name, epsg)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return

        QMessageBox.information(
            self, "Layer added",
            f"Wrote {len(points)} feature(s) to:\n{path}\n\n"
            f"Layer '{layer_name}' added at EPSG:{epsg}.",
        )

    @staticmethod
    def _write_gpkg_and_load(
        points: list[dict], path: str, layer_name: str, epsg: int,
    ) -> None:
        """Write the points (in WGS84) to a GeoPackage, reprojecting to
        the target EPSG, and add the resulting layer to the QGIS TOC."""
        from qgis.core import (
            QgsCoordinateReferenceSystem,
            QgsCoordinateTransform,
            QgsFeature,
            QgsField,
            QgsGeometry,
            QgsPointXY,
            QgsProject,
            QgsVectorFileWriter,
            QgsVectorLayer,
            QgsWkbTypes,
        )
        try:
            from qgis.PyQt.QtCore import QVariant
        except ImportError:
            QVariant = None  # Qt6 path uses Python types directly

        src_crs = QgsCoordinateReferenceSystem("EPSG:4326")
        dst_crs = QgsCoordinateReferenceSystem(f"EPSG:{epsg}")
        tr = QgsCoordinateTransform(src_crs, dst_crs, QgsProject.instance())

        # Build an in-memory layer first.
        mem = QgsVectorLayer(
            f"Point?crs=EPSG:{epsg}", layer_name, "memory",
        )
        prov = mem.dataProvider()
        if QVariant is not None:
            fields = [
                QgsField("name", QVariant.String),
                QgsField("lat_dd", QVariant.Double),
                QgsField("lon_dd", QVariant.Double),
            ]
        else:
            fields = [
                QgsField("name", str),
                QgsField("lat_dd", float),
                QgsField("lon_dd", float),
            ]
        prov.addAttributes(fields)
        mem.updateFields()

        feats = []
        for p in points:
            geom_4326 = QgsGeometry.fromPointXY(QgsPointXY(p["lon"], p["lat"]))
            geom_4326.transform(tr)
            f = QgsFeature(mem.fields())
            f.setGeometry(geom_4326)
            f.setAttributes([p["name"], float(p["lat"]), float(p["lon"])])
            feats.append(f)
        prov.addFeatures(feats)
        mem.updateExtents()

        # Write to GeoPackage. Append if file already exists.
        opts = QgsVectorFileWriter.SaveVectorOptions()
        opts.driverName = "GPKG"
        opts.layerName = layer_name
        if Path(path).exists():
            opts.actionOnExistingFile = (
                QgsVectorFileWriter.CreateOrOverwriteLayer
            )
        # writeAsVectorFormatV3 is the modern API; V2 is the fallback.
        ctx = QgsProject.instance().transformContext()
        if hasattr(QgsVectorFileWriter, "writeAsVectorFormatV3"):
            res = QgsVectorFileWriter.writeAsVectorFormatV3(
                mem, path, ctx, opts,
            )
        else:
            res = QgsVectorFileWriter.writeAsVectorFormatV2(
                mem, path, ctx, opts,
            )
        # writeAsVectorFormat* returns (errorCode, errorMessage[, newFilename, newLayer])
        err_code = res[0] if isinstance(res, tuple) else res
        if err_code != 0:
            err_msg = res[1] if isinstance(res, tuple) and len(res) > 1 else "unknown error"
            raise RuntimeError(f"GeoPackage write failed: {err_msg}")

        # Re-open from disk so QGIS gets the proper backing store.
        uri = f"{path}|layername={layer_name}"
        layer = QgsVectorLayer(uri, layer_name, "ogr")
        if not layer.isValid():
            raise RuntimeError(
                f"Wrote {path} but the layer '{layer_name}' did not load."
            )
        QgsProject.instance().addMapLayer(layer)
